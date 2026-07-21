"""Import a company's filled multi-sheet Excel workbook into the database.

Reads each sheet (Projects, Subcontractors, Subcontracts, Schedule_of_Values,
Progress_Claims, ... / Suppliers, Materials_Catalog, Requirements_BOQ,
Purchase_Orders, PO_Lines, Deliveries_GRN, Invoices), replaces the module's
tables, and lets the system derive everything else on the next read.
"""
from __future__ import annotations

import datetime as dt
import io

from openpyxl import load_workbook

from . import scoring
from .models import (
    ChangeOrder, Delivery, Invoice, Material, MaterialRequirement, POLine,
    ProgressClaim, Project, PurchaseOrder, SovLine, Subcontract,
    Subcontractor, Supplier,
)


def _s(v):
    return str(v).strip() if v not in (None, "") else None


def _n(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _i(v):
    try:
        return int(float(v)) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _d(v):
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return dt.date.fromisoformat(str(v)[:10])


def _read(wb, sheet, headers):
    """Return list of dicts from a sheet; finds the header row automatically."""
    if sheet not in wb.sheetnames:
        return []
    rows = list(wb[sheet].iter_rows(values_only=True))
    hidx, cmap = None, {}
    for i, row in enumerate(rows):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if headers[0] in vals:
            hidx = i
            cmap = {h: j for j, h in enumerate(vals) if h in headers}
            break
    if hidx is None:
        return []
    out = []
    for row in rows[hidx + 1:]:
        if row is None or all(c in (None, "") for c in row):
            continue
        j0 = cmap.get(headers[0])
        if j0 is None or j0 >= len(row) or row[j0] in (None, ""):
            continue
        out.append({h: (row[j] if j < len(row) else None) for h, j in cmap.items()})
    return out


def _upsert_projects(db, wb):
    for r in _read(wb, "Projects", ["project_id", "name", "client",
                                    "start_date", "planned_end_date", "status"]):
        pid = _s(r["project_id"])
        p = db.get(Project, pid) or Project(project_id=pid)
        p.name = _s(r.get("name")) or pid
        p.client = _s(r.get("client"))
        p.start_date = _d(r.get("start_date"))
        p.planned_end_date = _d(r.get("planned_end_date"))
        p.status = _s(r.get("status")) or "Active"
        db.add(p)


def import_subcontractor(content: bytes, db) -> dict:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    _upsert_projects(db, wb)
    for M in (ChangeOrder, ProgressClaim, SovLine, Subcontract, Subcontractor):
        db.query(M).delete()
    db.flush()

    counts = {}
    subs = _read(wb, "Subcontractors",
                 ["vendor_id", "company_name", "trade", "capacity_projects"])
    for r in subs:
        db.add(Subcontractor(vendor_id=_s(r["vendor_id"]),
                             company_name=_s(r.get("company_name")) or _s(r["vendor_id"]),
                             trade=_s(r.get("trade")),
                             capacity_projects=_i(r.get("capacity_projects"))))
    counts["subcontractors"] = len(subs)

    sc = _read(wb, "Subcontracts",
               ["subcontract_id", "vendor_id", "project_id", "title", "start_date",
                "planned_end_date", "status", "retainage_pct", "quality_score",
                "safety_score", "inspection_pass", "delay_days", "open_issues"])
    for r in sc:
        db.add(Subcontract(
            subcontract_id=_s(r["subcontract_id"]), vendor_id=_s(r.get("vendor_id")),
            project_id=_s(r.get("project_id")), title=_s(r.get("title")),
            start_date=_d(r.get("start_date")), planned_end_date=_d(r.get("planned_end_date")),
            status=_s(r.get("status")) or "Active", retainage_pct=_n(r.get("retainage_pct")),
            quality_score=_n(r.get("quality_score")), safety_score=_n(r.get("safety_score")),
            inspection_pass=_n(r.get("inspection_pass")), delay_days=_i(r.get("delay_days")),
            open_issues=_i(r.get("open_issues"))))
    counts["subcontracts"] = len(sc)

    sov = _read(wb, "Schedule_of_Values",
                ["line_id", "subcontract_id", "cost_code", "description",
                 "scheduled_value", "percent_complete"])
    for r in sov:
        db.add(SovLine(line_id=_s(r["line_id"]), subcontract_id=_s(r.get("subcontract_id")),
                       cost_code=_s(r.get("cost_code")), description=_s(r.get("description")),
                       scheduled_value=_n(r.get("scheduled_value")),
                       percent_complete=_n(r.get("percent_complete")) or 0))
    counts["sov_lines"] = len(sov)

    claims = _read(wb, "Progress_Claims",
                   ["claim_id", "subcontract_id", "period_end",
                    "percent_complete", "completed_value"])
    for r in claims:
        db.add(ProgressClaim(claim_id=_s(r["claim_id"]),
                             subcontract_id=_s(r.get("subcontract_id")),
                             period_end=_d(r.get("period_end")),
                             percent_complete=_n(r.get("percent_complete")),
                             completed_value=_n(r.get("completed_value"))))
    counts["claims"] = len(claims)

    cos = _read(wb, "Change_Orders",
                ["co_id", "subcontract_id", "description", "amount", "status", "co_date"])
    for r in cos:
        db.add(ChangeOrder(co_id=_s(r["co_id"]), subcontract_id=_s(r.get("subcontract_id")),
                           description=_s(r.get("description")), amount=_n(r.get("amount")),
                           status=_s(r.get("status")) or "Pending", co_date=_d(r.get("co_date"))))
    counts["change_orders"] = len(cos)

    # If inspection records are provided, compute quality/safety from them.
    insp = _read(wb, "Inspections_Safety",
                 ["subcontract_id", "period_end", "inspections_total", "inspections_passed",
                  "ncrs_raised", "ncrs_closed", "recordable_incidents", "man_hours"])
    for r in insp:
        s = db.get(Subcontract, _s(r["subcontract_id"]))
        if s is None:
            continue
        q = scoring.quality_and_inspection(r.get("inspections_total") or 0,
                                           r.get("inspections_passed") or 0,
                                           r.get("ncrs_raised") or 0, r.get("ncrs_closed") or 0)
        sf = scoring.safety(r.get("recordable_incidents") or 0, r.get("man_hours") or 0)
        if q["quality_score"] is not None:
            s.quality_score = q["quality_score"]
            s.inspection_pass = q["inspection_pass"]
        if sf["safety_score"] is not None:
            s.safety_score = sf["safety_score"]
    counts["inspections"] = len(insp)

    db.commit()
    return counts


def import_material(content: bytes, db) -> dict:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    _upsert_projects(db, wb)
    for M in (Invoice, Delivery, POLine, PurchaseOrder, MaterialRequirement,
              Material, Supplier):
        db.query(M).delete()
    db.flush()

    counts = {}
    sup = _read(wb, "Suppliers", ["supplier_id", "name", "category"])
    for r in sup:
        db.add(Supplier(supplier_id=_s(r["supplier_id"]), name=_s(r.get("name")),
                        category=_s(r.get("category"))))
    counts["suppliers"] = len(sup)

    mat = _read(wb, "Materials_Catalog", ["material_id", "name", "category", "unit"])
    for r in mat:
        db.add(Material(material_id=_s(r["material_id"]), name=_s(r.get("name")),
                        category=_s(r.get("category")), unit=_s(r.get("unit"))))
    counts["materials"] = len(mat)

    req = _read(wb, "Requirements_BOQ",
                ["req_id", "project_id", "material_id", "required_qty",
                 "minimum_stock", "consumed_qty"])
    for r in req:
        db.add(MaterialRequirement(
            req_id=_s(r["req_id"]), project_id=_s(r.get("project_id")),
            material_id=_s(r.get("material_id")), required_qty=_n(r.get("required_qty")),
            minimum_stock=_n(r.get("minimum_stock")), consumed_qty=_n(r.get("consumed_qty"))))
    counts["requirements"] = len(req)

    pos = _read(wb, "Purchase_Orders",
                ["po_id", "supplier_id", "project_id", "order_date",
                 "expected_delivery", "status"])
    for r in pos:
        db.add(PurchaseOrder(po_id=_s(r["po_id"]), supplier_id=_s(r.get("supplier_id")),
                             project_id=_s(r.get("project_id")), order_date=_d(r.get("order_date")),
                             expected_delivery=_d(r.get("expected_delivery")),
                             status=_s(r.get("status")) or "Open"))
    counts["purchase_orders"] = len(pos)

    lines = _read(wb, "PO_Lines",
                  ["line_id", "po_id", "material_id", "qty_ordered", "unit_price"])
    for r in lines:
        db.add(POLine(line_id=_s(r["line_id"]), po_id=_s(r.get("po_id")),
                      material_id=_s(r.get("material_id")), qty_ordered=_n(r.get("qty_ordered")),
                      unit_price=_n(r.get("unit_price"))))
    counts["po_lines"] = len(lines)

    dels = _read(wb, "Deliveries_GRN",
                 ["delivery_id", "po_id", "material_id", "project_id", "qty_received",
                  "qty_rejected", "order_date", "expected_date", "received_date"])
    for r in dels:
        db.add(Delivery(delivery_id=_s(r["delivery_id"]), po_id=_s(r.get("po_id")),
                        material_id=_s(r.get("material_id")), project_id=_s(r.get("project_id")),
                        qty_received=_n(r.get("qty_received")), qty_rejected=_n(r.get("qty_rejected")),
                        order_date=_d(r.get("order_date")), expected_date=_d(r.get("expected_date")),
                        received_date=_d(r.get("received_date"))))
    counts["deliveries"] = len(dels)

    inv = _read(wb, "Invoices",
                ["invoice_id", "po_id", "supplier_id", "invoice_date", "billed_qty", "amount"])
    for r in inv:
        db.add(Invoice(invoice_id=_s(r["invoice_id"]), po_id=_s(r.get("po_id")),
                       supplier_id=_s(r.get("supplier_id")), invoice_date=_d(r.get("invoice_date")),
                       billed_qty=_n(r.get("billed_qty")), amount=_n(r.get("amount")),
                       status="Pending"))
    counts["invoices"] = len(inv)

    db.commit()
    return counts
