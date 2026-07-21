"""Live-data ingestion: parse uploaded CSV/Excel and upsert into PostgreSQL.

Companies bring their data as spreadsheets in the same shape as the spec files
(a `Company_Input_Data` sheet). This module parses either CSV or .xlsx, maps the
columns by header name (tolerant of case, `%`, and `(days)` suffixes), coerces
types, and upserts rows by primary key.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import re
from typing import Any

from openpyxl import load_workbook

from .models import Material, Subcontractor

# Field spec per module: (model_field, display_header, type)
# type in {"str", "num", "int", "date"}. First entry is the primary key.
SUB_SPEC = [
    ("vendor_id", "Vendor ID", "str"),
    ("vendor_name", "Vendor Name", "str"),
    ("trade", "Trade", "str"),
    ("project", "Project", "str"),
    ("contract_value", "Contract Value", "num"),
    ("planned_progress", "Planned Progress %", "num"),
    ("actual_progress", "Actual Progress %", "num"),
    ("quality_score", "Quality Score", "num"),
    ("safety_score", "Safety Score", "num"),
    ("inspection_pass", "Inspection Pass %", "num"),
    ("delay_days", "Delay Days", "int"),
    ("open_issues", "Open Issues", "int"),
    ("invoice_amount", "Invoice Amount", "num"),
    ("paid_amount", "Paid Amount", "num"),
    ("engineer_rating", "Engineer Rating", "num"),
    ("client_rating", "Client Rating", "num"),
    ("active_projects", "Active Projects", "int"),
    ("capacity_projects", "Capacity Projects", "int"),
]

# One spreadsheet row = one assignment (subcontractor on a project). The upload
# endpoint upserts the company, the project, and the assignment from each row.
ASSIGN_SPEC = [
    ("assignment_id", "Assignment ID", "str"),
    ("vendor_id", "Company ID", "str"),
    ("company_name", "Company Name", "str"),
    ("trade", "Trade", "str"),
    ("capacity_projects", "Capacity Projects", "int"),
    ("project_id", "Project ID", "str"),
    ("project_name", "Project Name", "str"),
    ("contract_value", "Contract Value", "num"),
    ("planned_progress", "Planned Progress %", "num"),
    ("actual_progress", "Actual Progress %", "num"),
    ("quality_score", "Quality Score", "num"),
    ("safety_score", "Safety Score", "num"),
    ("inspection_pass", "Inspection Pass %", "num"),
    ("delay_days", "Delay Days", "int"),
    ("open_issues", "Open Issues", "int"),
    ("invoice_amount", "Invoice Amount", "num"),
    ("paid_amount", "Paid Amount", "num"),
]

MAT_SPEC = [
    ("material_id", "Material ID", "str"),
    ("material_name", "Material Name", "str"),
    ("category", "Category", "str"),
    ("current_stock", "Current Stock", "num"),
    ("minimum_stock", "Minimum Stock", "num"),
    ("required_qty", "Required Qty", "num"),
    ("supplier", "Supplier", "str"),
    ("lead_time_days", "Lead Time (days)", "int"),
    ("unit_price", "Unit Price", "num"),
    ("delivery_reliability", "Delivery Reliability %", "num"),
    ("project", "Project", "str"),
    ("expected_delivery", "Expected Delivery", "date"),
]

MODULES = {
    "subcontractors": {"model": Subcontractor, "spec": SUB_SPEC},
    "materials": {"model": Material, "spec": MAT_SPEC},
}


def _norm(header: Any) -> str:
    """Normalise a header for tolerant matching."""
    h = str(header or "").strip().lower()
    h = h.replace("%", " ")
    h = re.sub(r"\(days\)", " ", h)
    h = re.sub(r"[^a-z0-9]+", " ", h)
    return re.sub(r"\s+", " ", h).strip()


def _colmap(spec) -> dict[str, tuple[str, str]]:
    return {_norm(header): (field, typ) for field, header, typ in spec}


def coerce(value: Any, typ: str):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if typ == "str":
        return str(value).strip()
    if typ == "int":
        return int(round(float(value)))
    if typ == "num":
        return float(value)
    if typ == "date":
        if isinstance(value, (dt.date, dt.datetime)):
            return value.date() if isinstance(value, dt.datetime) else value
        return dt.date.fromisoformat(str(value).strip()[:10])
    return value


def coerce_record(rec: dict, spec) -> dict:
    """Coerce a JSON record (from the manual add form) to model-field types."""
    types = {field: typ for field, _, typ in spec}
    out: dict[str, Any] = {}
    for key, value in rec.items():
        if key in types and value not in (None, ""):
            out[key] = coerce(value, types[key])
    return out


def parse_table(content: bytes, filename: str) -> list[list[Any]]:
    """Return raw rows (including header) from a CSV or Excel file."""
    name = (filename or "").lower()
    if name.endswith(".csv") or (not name.endswith((".xlsx", ".xlsm"))
                                 and b"," in content[:2000]):
        text = content.decode("utf-8-sig", errors="replace")
        return [row for row in csv.reader(io.StringIO(text))]
    if name.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb["Company_Input_Data"] if "Company_Input_Data" in wb.sheetnames \
            else wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    raise ValueError("Unsupported file type — please upload .csv or .xlsx")


def rows_to_records(rows: list[list[Any]], spec) -> tuple[list[dict], list[str]]:
    """Map raw rows to model-field dicts. Returns (records, warnings)."""
    if not rows:
        return [], ["File is empty"]
    colmap = _colmap(spec)
    pk_field = spec[0][0]
    header = rows[0]
    idx_field: dict[int, tuple[str, str]] = {}
    for i, h in enumerate(header):
        hit = colmap.get(_norm(h))
        if hit:
            idx_field[i] = hit

    if pk_field not in {f for f, _ in idx_field.values()}:
        return [], [f"Required column '{spec[0][1]}' not found in the file"]

    records, warnings = [], []
    for n, row in enumerate(rows[1:], start=2):
        if not any(c not in (None, "") for c in row):
            continue
        rec: dict[str, Any] = {}
        for i, (field, typ) in idx_field.items():
            if i < len(row):
                try:
                    val = coerce(row[i], typ)
                    if val is not None:
                        rec[field] = val
                except (ValueError, TypeError):
                    warnings.append(f"Row {n}: bad value for '{field}'")
        if rec.get(pk_field):
            records.append(rec)
        else:
            warnings.append(f"Row {n}: missing {spec[0][1]}, skipped")
    return records, warnings


def upsert(db, model, spec, records: list[dict], replace: bool = False) -> dict:
    """Insert new rows / update existing rows by primary key."""
    pk_field = spec[0][0]
    if replace:
        db.query(model).delete()
        db.flush()
    inserted = updated = 0
    for rec in records:
        pk = rec.get(pk_field)
        obj = None if replace else db.get(model, pk)
        if obj:
            for k, v in rec.items():
                setattr(obj, k, v)
            updated += 1
        else:
            db.add(model(**rec))
            inserted += 1
    db.commit()
    return {"inserted": inserted, "updated": updated, "total": len(records)}


def template_csv(spec) -> str:
    """A header-only CSV (plus one example row) for the module."""
    headers = [header for _, header, _ in spec]
    example = {
        "str": "Example", "num": "0", "int": "0", "date": "2026-07-15",
    }
    row = [example[typ] for _, _, typ in spec]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerow(row)
    return buf.getvalue()
